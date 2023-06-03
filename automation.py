import csv
import os
import time
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import tabula
import pandas as pd
import random
import camelot

# Create the 'Pdf Files' folder if it doesn't exist
pdf_folder = 'Pdf Files'
if not os.path.exists(pdf_folder):
    os.makedirs(pdf_folder)

# Configure Chrome options for WebDriver
chrome_options = Options()
chrome_options.add_argument("--headless")  # Run Chrome in headless mode (no GUI)
chrome_options.add_argument("--disable-dev-shm-usage")  # Disable /dev/shm usage
chrome_options.add_argument("--no-sandbox")  # Disable sandbox mode

# Initialize WebDriver
driver = webdriver.Chrome(options=chrome_options)


def perform_google_search(keywords):
    url_list = []
    try:
        # Open Google search in WebDriver
        driver.get("https://www.google.com")
        search_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "q")))

        # Enter keywords in search box and submit
        search_box.send_keys(keywords)
        search_box.send_keys(Keys.RETURN)

        # Wait for the search results to load
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "search")))

        # Extract URLs from the search results
        search_results = driver.find_elements(By.XPATH, "//div[@class='yuRUbf']/a")
        for result in search_results:
            url = result.get_attribute("href")
            url_list.append(url)

    except Exception as e:
        print("Error performing Google search:", e)

    return url_list


def extract_text_data(url_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Set column names
    ws["A1"] = "URL"
    ws["B1"] = "Title"
    ws["C1"] = "Text"

    # Remove common irrelevant data from websites
    irrelevant_data = ["cookie", "advertisement", "privacy policy"]

    row_index = 2  # Start from row 2

    for url in url_list:
        try:
            # Open URL in WebDriver
            driver.get(url)

            # Extract relevant data
            soup = BeautifulSoup(driver.page_source, "html.parser")
            title = soup.title.string.strip()
            text = soup.get_text()

            # Remove irrelevant data from text
            for phrase in irrelevant_data:
                text = text.replace(phrase, "")

            # Save data to Excel
            ws.cell(row=row_index, column=1, value=url)
            ws.cell(row=row_index, column=2, value=title)
            ws.cell(row=row_index, column=3, value=text.strip())

            row_index += 1

        except Exception as e:
            print("Error extracting data from URL:", url)
            print(e)

    # Auto-fit column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.coordinate == f"{column_letter}1":
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the Excel file
    output_filename = "output.xlsx"
    wb.save(output_filename)
    print("Data extraction complete. Output file:", output_filename)


def download_pdf_files(url_list):
    for url in url_list:
        try:
            response = requests.get(url)
            if response.headers.get("Content-Type") == "application/pdf":
                # Create 'Pdf Files' folder if it doesn't exist
                if not os.path.exists(pdf_folder):
                    os.makedirs(pdf_folder)

                # Extract filename from URL
                filename = url.split("/")[-1]

                # Download the PDF file
                with open(os.path.join(pdf_folder, filename), "wb") as file:
                    file.write(response.content)
                print("Downloaded:", filename)

        except Exception as e:
            print("Error downloading PDF from URL:", url)
            print(e)


def extract_tables_from_pdf(input_pdf, output_excel):
    # # Read PDF tables using tabula
    # pdf_tables = tabula.read_pdf(input_pdf, pages='all', multiple_tables=True)
    #
    # # Create an Excel writer
    # excel_writer = pd.ExcelWriter(output_excel)
    #
    # # Iterate over each table and save as separate sheets in Excel
    # for i, table in enumerate(pdf_tables, start=1):
    #     sheet_name = f"Table {i}"
    #     table.to_excel(excel_writer, sheet_name=sheet_name, index=False)
    #
    # # Save the Excel file
    # excel_writer.save()
    tables = camelot.read_pdf(input_pdf, flavor='stream', pages='all')
    print(f'{tables.n} no of tables found in the file: {input_pdf}')
    if tables.n > 0:
        writer = pd.ExcelWriter(f'{output_excel.strip(".pdf")}.xlsx', engine='xlsxwriter')
        for i, table in enumerate(tables):
            table.df.to_excel(writer, sheet_name=f'Table_{i + 1}', index=False)
        writer.save()
        print("Table extraction complete. Output file:{}.xlsx".format(output_excel.strip(".pdf")))
    else:
        print("No tables found in {} file".format(input_pdf))


def extract_tables_from_pdf_files():
    pdf_files = os.listdir(pdf_folder)
    wb = Workbook()

    for filename in pdf_files:
        try:
            extract_tables_from_pdf(filename, filename)

        except Exception as e:
            print("Error extracting tables from PDF:", filename)
            print(e)


if __name__ == '__main__':
    keywords = input("Enter keywords for Google search: ")
    url_list = perform_google_search(keywords)

    if url_list:
        extract_text_data(url_list)
        download_pdf_files(url_list)
        extract_tables_from_pdf_files()

    # Clean up resources
    driver.quit()
